"""Searchable index of weather stations from climate.onebuilding.org."""

from __future__ import annotations

import gzip
import json
import math
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from .spatial import haversine_km
from .station import SearchResult, SpatialResult, WeatherStation

# The main regional TMYx Excel index files covering worldwide stations.
_INDEX_FILES: tuple[str, ...] = (
    "Region1_Africa_TMYx_EPW_Processing_locations.xlsx",
    "Region2_Asia_TMYx_EPW_Processing_locations.xlsx",
    "Region2_Region6_Russia_TMYx_EPW_Processing_locations.xlsx",
    "Region3_South_America_TMYx_EPW_Processing_locations.xlsx",
    "Region4_USA_TMYx_EPW_Processing_locations.xlsx",
    "Region4_Canada_TMYx_EPW_Processing_locations.xlsx",
    "Region4_NA_CA_Caribbean_TMYx_EPW_Processing_locations.xlsx",
    "Region5_Southwest_Pacific_TMYx_EPW_Processing_locations.xlsx",
    "Region6_Europe_TMYx_EPW_Processing_locations.xlsx",
    "Region7_Antarctica_TMYx_EPW_Processing_locations.xlsx",
)

_SOURCES_BASE_URL = "https://climate.onebuilding.org/sources"
_USER_AGENT = "idfkit (https://github.com/samuelduchesne/idfkit)"

_BUNDLED_INDEX = Path(__file__).parent / "data" / "stations.json.gz"
_CACHED_INDEX = "stations.json.gz"


def default_cache_dir() -> Path:
    """Return the platform-appropriate cache directory for idfkit weather data."""
    if sys.platform == "win32":
        base = Path(os.environ.get("LOCALAPPDATA", Path.home() / "AppData" / "Local"))
        return base / "idfkit" / "cache" / "weather"
    if sys.platform == "darwin":
        return Path.home() / "Library" / "Caches" / "idfkit" / "weather"
    # Linux / other POSIX
    xdg = os.environ.get("XDG_CACHE_HOME")
    base = Path(xdg) if xdg else Path.home() / ".cache"
    return base / "idfkit" / "weather"


# ---------------------------------------------------------------------------
# Download / parse helpers (used by refresh and build script)
# ---------------------------------------------------------------------------


def _download_file(url: str, dest: Path) -> str | None:
    """Download a file from *url* to *dest*, creating parent dirs as needed.

    Returns the ``Last-Modified`` response header value, or ``None`` if
    the header is absent.
    """
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = Request(url, headers={"User-Agent": _USER_AGENT})  # noqa: S310
    with urlopen(req, timeout=60) as resp:  # noqa: S310
        last_modified: str | None = resp.headers.get("Last-Modified")
        dest.write_bytes(resp.read())
    return last_modified


def _ensure_index_file(filename: str, cache_dir: Path) -> tuple[Path, str | None]:
    """Return the local path for an Excel index file, downloading if absent.

    Returns ``(path, last_modified_header)``.  When the file already exists
    in the cache the header is ``None`` (we don't know it).
    """
    local = cache_dir / "indexes" / filename
    if local.exists():
        return local, None
    url = f"{_SOURCES_BASE_URL}/{filename}"
    try:
        last_modified = _download_file(url, local)
    except (HTTPError, URLError, TimeoutError, OSError) as exc:
        msg = f"Failed to download weather index {filename}: {exc}"
        raise RuntimeError(msg) from exc
    return local, last_modified


def _parse_excel(path: Path) -> list[WeatherStation]:
    """Parse a single climate.onebuilding.org Excel index file.

    Columns (1-indexed):
        A: Country, B: State, C: City/Station, D: WMO, E: Source Data,
        F: Latitude (N+/S-), G: Longitude (E+/W-), H: Time Zone (GMT +/-),
        I: Elevation (m), J: URL
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        msg = (
            "openpyxl is required for refreshing the weather station index. "
            "Install it with:  pip install idfkit[weather]"
        )
        raise ImportError(msg)  # noqa: B904

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    stations: list[WeatherStation] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # Skip header row
            continue
        # Unpack the 10 columns.  Cast via Any because openpyxl's
        # cell-value union type is too wide for int()/float() directly.
        vals: tuple[Any, ...] = tuple(row)
        country, state, city, wmo, source, lat, lon, tz, elev, url = (
            vals[0],
            vals[1],
            vals[2],
            vals[3],
            vals[4],
            vals[5],
            vals[6],
            vals[7],
            vals[8],
            vals[9],
        )
        if url is None or lat is None or lon is None:
            continue
        stations.append(
            WeatherStation(
                country=str(country or ""),
                state=str(state or ""),
                city=str(city or ""),
                wmo=str(int(wmo)) if wmo is not None and str(wmo).replace(".", "").isdigit() else "",
                source=str(source or ""),
                latitude=float(lat),
                longitude=float(lon),
                timezone=float(tz) if tz is not None else 0.0,
                elevation=float(elev) if elev is not None else 0.0,
                url=str(url),
            )
        )
    wb.close()
    return stations


# ---------------------------------------------------------------------------
# Compressed index serialization
# ---------------------------------------------------------------------------


def _load_compressed_index(path: Path) -> tuple[list[WeatherStation], dict[str, str], str]:
    """Load a gzip-compressed JSON station index.

    Returns ``(stations, last_modified_headers, built_at_iso)``.
    """
    with gzip.open(path, "rt", encoding="utf-8") as f:
        data: dict[str, Any] = json.load(f)
    stations = [WeatherStation.from_dict(d) for d in data["stations"]]
    last_modified: dict[str, str] = data.get("last_modified", {})
    built_at: str = data.get("built_at", "")
    return stations, last_modified, built_at


def _save_compressed_index(
    stations: list[WeatherStation],
    last_modified: dict[str, str],
    dest: Path,
) -> None:
    """Serialize stations and metadata to a gzip-compressed JSON file."""
    data = {
        "built_at": datetime.now(tz=timezone.utc).isoformat(),
        "last_modified": last_modified,
        "stations": [s.to_dict() for s in stations],
    }
    dest.parent.mkdir(parents=True, exist_ok=True)
    with gzip.open(dest, "wt", encoding="utf-8") as f:
        json.dump(data, f, separators=(",", ":"))


def _head_last_modified(url: str) -> str | None:
    """Send a HEAD request and return the ``Last-Modified`` header, or ``None``."""
    req = Request(url, method="HEAD", headers={"User-Agent": _USER_AGENT})  # noqa: S310
    try:
        with urlopen(req, timeout=30) as resp:  # noqa: S310
            return resp.headers.get("Last-Modified")
    except (HTTPError, URLError, TimeoutError, OSError):
        return None


# ---------------------------------------------------------------------------
# Fuzzy search scoring
# ---------------------------------------------------------------------------


def _score_station(station: WeatherStation, query: str, tokens: list[str]) -> tuple[float, str]:
    """Score a station against a search query.

    Returns ``(score, match_field)`` where score is 0.0-1.0.
    """
    name_lower = station.city.lower().replace(".", " ").replace("-", " ")
    display_lower = station.display_name.lower()

    # Signal 1: Exact WMO match (compare as strings, stripping leading zeros for flexibility)
    if query.isdigit() and query.lstrip("0") == station.wmo.lstrip("0"):
        return 1.0, "wmo"

    # Signal 2: Full query is a substring of the display name
    if query in display_lower:
        coverage = len(query) / max(len(display_lower), 1)
        return 0.85 + 0.1 * coverage, "name"

    # Signal 3: Full query is a substring of the city name (dots -> spaces)
    if query in name_lower:
        coverage = len(query) / max(len(name_lower), 1)
        return 0.85 + 0.1 * coverage, "name"

    # Signal 4: All query tokens appear in the name
    name_tokens = set(name_lower.split())
    if tokens and all(any(t.startswith(qt) for t in name_tokens) for qt in tokens):
        coverage = sum(len(qt) for qt in tokens) / max(len(name_lower), 1)
        return 0.6 + 0.3 * min(coverage, 1.0), "name"

    # Signal 5: Partial token overlap (prefix matching)
    if tokens:
        matching = sum(1 for qt in tokens if any(t.startswith(qt) for t in name_tokens))
        if matching > 0:
            ratio = matching / len(tokens)
            return 0.3 * ratio, "name"

    # Signal 6: State or country match
    if query == station.state.lower():
        return 0.5, "state"
    if query == station.country.lower():
        return 0.4, "country"

    return 0.0, ""


# ---------------------------------------------------------------------------
# StationIndex
# ---------------------------------------------------------------------------


class StationIndex:
    """Searchable index of weather stations from climate.onebuilding.org.

    Use :meth:`load` to load the bundled (or user-refreshed) station index.
    No network access or ``openpyxl`` is required for :meth:`load`.

    Use :meth:`check_for_updates` to see if upstream data has changed, and
    :meth:`refresh` to re-download and rebuild the index.

    Example::

        index = StationIndex.load()
        results = index.search("chicago ohare", limit=3)
        for r in results:
            print(r.station.display_name, r.score)
    """

    __slots__ = ("_by_wmo", "_last_modified", "_stations")

    _stations: list[WeatherStation]
    _by_wmo: dict[str, list[WeatherStation]]
    _last_modified: dict[str, str]

    def __init__(self, stations: list[WeatherStation]) -> None:
        self._stations = stations
        self._by_wmo: dict[str, list[WeatherStation]] = {}
        for s in stations:
            self._by_wmo.setdefault(s.wmo, []).append(s)
        self._last_modified: dict[str, str] = {}

    # --- Construction -------------------------------------------------------

    @classmethod
    def load(cls, *, cache_dir: Path | None = None) -> StationIndex:
        """Load the station index from a local compressed file.

        Checks for a user-refreshed cache first, then falls back to the
        bundled index shipped with the package.  No network access is
        required.

        Args:
            cache_dir: Override the default cache directory.
        """
        cache = cache_dir or default_cache_dir()
        cached_path = cache / _CACHED_INDEX

        if cached_path.is_file():
            source = cached_path
        elif _BUNDLED_INDEX.is_file():
            source = _BUNDLED_INDEX
        else:
            msg = (
                "No station index found. The bundled index is missing and no "
                "cached index exists. Run StationIndex.refresh() to download one."
            )
            raise FileNotFoundError(msg)

        stations, last_modified, _ = _load_compressed_index(source)
        instance = cls(stations)
        instance._last_modified = last_modified
        return instance

    @classmethod
    def from_stations(cls, stations: list[WeatherStation]) -> StationIndex:
        """Create an index from an explicit list of stations (useful for tests)."""
        return cls(stations)

    @classmethod
    def refresh(cls, *, cache_dir: Path | None = None) -> StationIndex:
        """Re-download Excel indexes from climate.onebuilding.org and rebuild the cache.

        Requires ``openpyxl``.  Install with ``pip install idfkit[weather]``.

        Args:
            cache_dir: Override the default cache directory.
        """
        cache = cache_dir or default_cache_dir()

        all_stations: list[WeatherStation] = []
        last_modified: dict[str, str] = {}
        for fname in _INDEX_FILES:
            local_path, lm = _ensure_index_file(fname, cache)
            if lm is not None:
                last_modified[fname] = lm
            all_stations.extend(_parse_excel(local_path))

        dest = cache / _CACHED_INDEX
        _save_compressed_index(all_stations, last_modified, dest)

        instance = cls(all_stations)
        instance._last_modified = last_modified
        return instance

    # --- Freshness ----------------------------------------------------------

    def check_for_updates(self) -> bool:
        """Check if upstream Excel files have changed since this index was built.

        Sends lightweight HEAD requests to climate.onebuilding.org.
        Returns ``True`` if any file has a newer ``Last-Modified`` date.
        Returns ``False`` if all files match or if the check fails (offline,
        timeout, etc.).
        """
        if not self._last_modified:
            return False
        for fname in _INDEX_FILES:
            stored = self._last_modified.get(fname)
            if stored is None:
                continue
            url = f"{_SOURCES_BASE_URL}/{fname}"
            upstream = _head_last_modified(url)
            if upstream is not None and upstream != stored:
                return True
        return False

    # --- Properties ---------------------------------------------------------

    @property
    def stations(self) -> list[WeatherStation]:
        """All stations in the index."""
        return list(self._stations)

    def __len__(self) -> int:
        return len(self._stations)

    # --- Exact lookups ------------------------------------------------------

    def get_by_wmo(self, wmo: str) -> list[WeatherStation]:
        """Look up stations by WMO number.

        Args:
            wmo: WMO station number as a string (e.g. ``"722950"``).

        Returns a list because a single WMO number can correspond to
        multiple stations or dataset variants.
        """
        return list(self._by_wmo.get(wmo, []))

    # --- Fuzzy text search --------------------------------------------------

    def search(
        self,
        query: str,
        *,
        limit: int = 10,
        country: str | None = None,
    ) -> list[SearchResult]:
        """Fuzzy-search stations by name, city, state, or WMO number.

        Matching is case-insensitive and uses substring / token-prefix
        heuristics (no external NLP dependencies).

        Args:
            query: Free-text search query.
            limit: Maximum number of results to return.
            country: If given, restrict to stations in this country code.
        """
        q = query.strip().lower()
        if not q:
            return []
        tokens = q.split()

        scored: list[SearchResult] = []
        for station in self._stations:
            if country and station.country.upper() != country.upper():
                continue
            score, match_field = _score_station(station, q, tokens)
            if score > 0:
                scored.append(SearchResult(station=station, score=score, match_field=match_field))

        scored.sort(key=lambda r: r.score, reverse=True)
        return scored[:limit]

    # --- Spatial search -----------------------------------------------------

    def nearest(
        self,
        latitude: float,
        longitude: float,
        *,
        limit: int = 5,
        max_distance_km: float | None = None,
        country: str | None = None,
    ) -> list[SpatialResult]:
        """Find stations nearest to a geographic coordinate.

        Uses the Haversine formula for great-circle distance.  A bounding-box
        pre-filter is applied when *max_distance_km* is specified to avoid
        computing distances for stations that are obviously too far.

        Args:
            latitude: Decimal degrees, north positive.
            longitude: Decimal degrees, east positive.
            limit: Maximum results to return.
            max_distance_km: Exclude stations farther than this.
            country: If given, restrict to this country code.
        """
        # Bounding-box pre-filter (~111 km per degree of latitude)
        if max_distance_km is not None:
            delta_deg = max_distance_km / 111.0 + 1.0  # small margin
            lat_min = latitude - delta_deg
            lat_max = latitude + delta_deg
            # Longitude degrees vary with latitude
            cos_lat = math.cos(math.radians(latitude))
            lon_delta = delta_deg / max(cos_lat, 0.01)
            lon_min = longitude - lon_delta
            lon_max = longitude + lon_delta
        else:
            lat_min = lat_max = lon_min = lon_max = 0.0  # unused

        results: list[SpatialResult] = []
        for station in self._stations:
            if country and station.country.upper() != country.upper():
                continue
            if max_distance_km is not None:
                if station.latitude < lat_min or station.latitude > lat_max:
                    continue
                if station.longitude < lon_min or station.longitude > lon_max:
                    continue
            dist = haversine_km(latitude, longitude, station.latitude, station.longitude)
            if max_distance_km is not None and dist > max_distance_km:
                continue
            results.append(SpatialResult(station=station, distance_km=dist))

        results.sort(key=lambda r: r.distance_km)
        return results[:limit]

    # --- Filtering ----------------------------------------------------------

    def filter(
        self,
        *,
        country: str | None = None,
        state: str | None = None,
        wmo_region: int | None = None,
    ) -> list[WeatherStation]:
        """Filter stations by metadata criteria.

        All specified criteria must match (logical AND).
        """
        result: list[WeatherStation] = []
        for s in self._stations:
            if country and s.country.upper() != country.upper():
                continue
            if state and s.state.upper() != state.upper():
                continue
            if wmo_region is not None:
                # Infer WMO region from the URL path
                url_lower = s.url.lower()
                if f"wmo_region_{wmo_region}" not in url_lower:
                    continue
            result.append(s)
        return result

    @property
    def countries(self) -> list[str]:
        """Sorted list of unique country codes in the index."""
        return sorted({s.country for s in self._stations})
